import requests
import json
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime

class Wrapper():

    def __init__(self):
        self.url = "https://api.leboncoin.fr/finder/search"
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/76.0.3809.100 Safari/537.36',
            'Accept-Language': '*',
            'Accept': '*/*',
            'Accept-Encoding':  'gzip, deflate, br',
            'Accept-Language': 'en-GB,en-US;q=0.9,en;q=0.8',
            'Content-Type': 'application/json',
            'Referer': "https://www.leboncoin.fr/recherche/",
            'Origin': 'https://www.leboncoin.fr'
        }
    def get_research(self, category, keyword):
        payload = {
            "limit":200,
            "filters":{
                "category":{"id":category},
                "enums":{ "ad_type": [ "offer" ] },
                "location":{ "region": "22", "department": "69", "city_zipcodes": [
                    { "zipcode": "69001" },
                    { "zipcode": "69002" },
                    { "zipcode": "69003" },
                    { "zipcode": "69004" },
                    { "zipcode": "69005" },
                    { "zipcode": "69006" },
                    { "zipcode": "69007" },
                    { "zipcode": "69008" },
                    { "zipcode": "69009" }]
                },
                "keywords":{ "text": keyword },
                "range":{}
            },
            "offset":0,
            "owner_type":"pro",
            "sort_by": None,
          "sort_order": None
        }
        resp = requests.post(self.url, data=json.dumps(payload), headers=self.headers)
        return resp.json()

    def export_to_excel(self, response,filename):
        workbook = load_workbook(filename="leboncoin.xlsx")
        sheet = workbook["Datas"]
        keys = response[0].keys()
        column_counter = 1
        row_counter = 1
        for row in response:
            if row_counter == 1:
                for key in keys:
                    sheet.cell(row = 1, column = column_counter).value = key
                    column_counter += 1
            else:
                column_counter =1
                for key, value in row.items():
                    print(key," : ",value)
                    if key == "price":
                        sheet.cell(row=row_counter, column=column_counter).value = value[0]
                    elif key == "images" or key == "attributes" or key == "location" or key == "owner" or key == "options" or key == "calendar":
                        pass
                    else:
                        sheet.cell(row=row_counter, column=column_counter).value = value
                    column_counter += 1
            row_counter += 1
        workbook.save(filename="leboncoin.xlsx")
        """with open(filename, "w") as outfile:
            writer = csv.writer(outfile)
            writer.writerow(response[0].keys())
            for row in response:
                writer.writerow(row.values())"""

    def print_item(self, object):
        print("Titre : "+ object["subject"])
        print("Date : ", datetime.strptime(object["first_publication_date"],"%Y-%m-%d %H:%M:%S"))
        if "price" in object:
            print("Prix : "+ str(object["price"][0]))
        print("Prix : "+ object["url"])
        print("Description : " + object["body"])
